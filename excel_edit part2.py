import openpyxl
from openpyxl.styles.numbers import BUILTIN_FORMATS
import string
import smtplib
from email.mime.text import MIMEText

abc = string.ascii_uppercase
wb = openpyxl.load_workbook('C:\py_project\pythonProject\Chromedriver\export.xlsx')
sheet = wb['Sheet1']
row_count = sheet.max_row
col_count = sheet.max_column
#Открываем EXCEL и заполняем столбец "Изменение"
for i in range(2, row_count + 1):
    sheet[f'K{i}'] = '-' if '-' in [sheet[f'I{i}'].value, sheet[f'D{i}'].value] else f'=I{i}/D{i}'
#Меняем формат ячеек"
for a in ["B", "D", "G", "I"]:
    for i in range(2, row_count + 1):
        tmp = '-' if sheet[f'{a}{i}'].value == '-' else int(sheet[f'{a}{i}'].value) / 10000
        sheet[f'{a}{i}'] = tmp
        sheet[f'{a}{i}'].number_format = '"₽"#,##0.0000_);("₽"#,##0.0000)'
print("Меняем формат ячеек")
print(BUILTIN_FORMATS[7])
#Меняем ширину столбцов"
for ind, col in enumerate(sheet.columns):
    max_len = 0
    name_col = abc[ind]
    for cell in col:
        if len(str(cell.value)) > max_len:
            max_len = len(str(cell.value))
        sheet.column_dimensions[str(name_col)].width = (max_len + 1) * 1.1

wb.save('moex_export_edit.xlsx')
print("Сохранили EXCEL")

#Формируем сообещние о количестве строк со склоняемым словом"

if row_count % 10 == 1 and row_count % 100 != 11:
    text = "строка"
elif row_count % 10 in [2, 3, 4] and row_count % 100 not in [12, 13, 14]:
    text = "строки"
else:
    text = "строк"
print(f' в файле {row_count} {text}')

#Отправляем на почту

sender = "yurymiheev@yandex.ru"
password = 'jtdobkrxstpdiwjo'   #разовый удаляемый пароль
server = smtplib.SMTP("smtp.yandex.ru", 465)
server.starttls()
msg = MIMEText()
server.login(sender, password)

msg.attach(MIMEText(f' в файле {row_count} {text}'))

with open("moex_export_edit.xlsx") as f:
    file = MIMEText(f.read())
    file.add_header('content-disposition', 'attachment', filename='')
    msg.attach(file)

server.sendmail(sender, sender, msg.as_string())
